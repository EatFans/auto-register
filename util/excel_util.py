# Excel表格处理工具
from openpyxl.workbook import Workbook
from entity.account import Account
from typing import List

def export_accounts_to_excel(accounts: List[Account], filepath: str):
    """
    导出账号数据excel表格
    :param accounts: 账号数据
    :param filepath: 导出文件路径
    :return: 如果导出保存成功就返回true，否则就返回false
    """
    try:
        # 创建 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Accounts"
        # 写入表头
        headers = ['邮箱','密码','名字','生日','国家区域','性别']
        ws.append(headers)
        # 写入每一行数据
        for account in accounts:
            # 处理生日字段，支持字符串和datetime对象
            if isinstance(account.birthday, str):
                birthday_str = account.birthday
            else:
                birthday_str = account.birthday.strftime('%Y-%m-%d')
            
            ws.append([
                account.email,
                account.password,
                account.name,
                birthday_str,
                account.country,
                account.gender
            ])
        # 保存文件到指定路径
        wb.save(filepath)
        return True
    except Exception as e:
        print(e)
        return False

def export_failed_data_to_excel(failed_data: List[dict], filepath: str):
    """
    导出失败的注册数据到Excel表格
    :param failed_data: 失败的注册数据列表
    :param filepath: 导出文件路径
    :return: 如果导出保存成功就返回true，否则就返回false
    """
    try:
        # 创建 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Failed Registrations"
        
        if not failed_data:
            return False
        
        # 获取所有可能的列名
        all_columns = set()
        for data in failed_data:
            all_columns.update(data.keys())
        
        # 排序列名，确保失败原因在最后
        columns = sorted([col for col in all_columns if col != '失败原因'])
        if '失败原因' in all_columns:
            columns.append('失败原因')
        
        # 写入表头
        ws.append(columns)
        
        # 写入每一行数据
        for data in failed_data:
            row = []
            for col in columns:
                value = data.get(col, '')
                # 处理生日字段
                if col == '生日' and hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                row.append(value)
            ws.append(row)
        
        # 保存文件到指定路径
        wb.save(filepath)
        return True
    except Exception as e:
        print(e)
        return False

def export_all_registration_data_to_excel(accounts: List[Account], failed_data: List[dict], filepath: str):
    """
    导出所有注册数据到一个Excel表格，包括成功和失败的数据
    :param accounts: 成功注册的账号数据
    :param failed_data: 失败的注册数据列表
    :param filepath: 导出文件路径
    :return: 如果导出保存成功就返回true，否则就返回false
    """
    try:
        # 创建 Excel 工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "Registration Results"
        
        # 写入表头
        headers = ['邮箱', '密码', '名字', '生日', '国家区域', '性别', '注册状态', '失败原因']
        ws.append(headers)
        
        # 写入成功注册的数据
        for account in accounts:
            # 处理生日字段，支持字符串和datetime对象
            if isinstance(account.birthday, str):
                birthday_str = account.birthday
            else:
                birthday_str = account.birthday.strftime('%Y-%m-%d')
            
            ws.append([
                account.email,
                account.password,
                account.name,
                birthday_str,
                account.country,
                account.gender,
                '成功',
                ''
            ])
        
        # 写入失败注册的数据
        for data in failed_data:
            # 处理生日字段
            birthday = data.get('生日', '')
            if hasattr(birthday, 'strftime'):
                birthday = birthday.strftime('%Y-%m-%d')
            
            # 对于失败的注册，只显示从Excel中读取到的原始信息
            # 不显示生成的邮箱等信息
            ws.append([
                '',  # 失败的注册不显示生成的邮箱
                '',  # 失败的注册密码为空
                data.get('名字', ''),
                birthday,
                data.get('国家区域', ''),
                data.get('性别', ''),
                '失败',
                data.get('失败原因', '')
            ])
        
        # 保存文件到指定路径
        wb.save(filepath)
        return True
    except Exception as e:
        print(e)
        return False
